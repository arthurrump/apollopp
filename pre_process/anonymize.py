import argparse
import base64
import csv
import json
import random
import re
import sys
import unicodedata
import uuid
import zipfile

from openpyxl import Workbook, load_workbook
from os import path, remove
from typing import Callable, Dict, Iterable, List, Mapping, Tuple, TypeVar, Union
from zipfile import ZipFile

T = TypeVar("T")
Key = TypeVar("Key")
Value = TypeVar("Value")

Replacements = Mapping[re.Pattern, Union[str, Callable[[re.Match[str]], str]]]

# Generate a random string to replace names
def generateRandomName() -> str:
    # Make sure that the generated ID is a valid variable name, in case someone
    # uses their own name to name a variable (you never know...)
    return "i" + base64.b64encode(uuid.uuid4().bytes, b"__").decode("utf-8").rstrip("=")

# Generate a random number to replace student numbers
def generateRandomId() -> str:
    # Random replacement for student numbers
    return str(random.randint(1000000, 9999999))

# Generate a random email to replace emails
def generateRandomEmail() -> str:
    return generateRandomName() + "@example.com"

# Remove accents to create ASCII-only strings, because people may write
# their own names without accents in filenames and code
# https://stackoverflow.com/a/517974
def removeAccents(input_str: str) -> str:
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    return u"".join([c for c in nfkd_form if not unicodedata.combining(c)])

# Make a shuffled copy of a list
def shuffled(l: List[T]) -> List[T]:
    lcopy = l.copy()
    random.shuffle(lcopy)
    return lcopy

# Replace all occurences of keys with their corresponding values
def replaceAll(s: str, replacements: Replacements) -> str:
    for pidPattern, replacement in replacements.items():
        s = pidPattern.sub(replacement, s)
    return s

# Replace every occurence of an email with a new random email from the dictionary
def replaceEmails(s: str) -> str:
    return re.sub(
        r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", 
        lambda _: generateRandomEmail(),
        s
    )

# Replace strings nested in objects and lists, like when dealing with data from
# a JSON structure
def replaceInObject(value: T, replacements: Replacements) -> T:
    if isinstance(value, dict):
        return { k: replaceInObject(v, replacements) for k, v in value.items() }
    elif isinstance(value, list):
        return [ replaceInObject(v, replacements) for v in value ]
    elif isinstance(value, str):
        return replaceAll(value, replacements)
    else:
        return value

# Helper to compile a case-insensitive regex
def regexi(s: str) -> re.Pattern:
    return re.compile(s, re.IGNORECASE)

# Helper to compile a case-insensitive regex that matches the exact string
def regexesci(s: str) -> re.Pattern:
    return regexi(re.escape(s))

# Create a map from patterns matching a name or identifier to a random alias
def generatePiiReplacements(students: Iterable[Tuple[str, str]]) -> Replacements:
    # Names to skip, because they are also rather important words in code
    skipNames = [ "max", "floor" ]
    map: Replacements = {}
    for name, id in students:
        # Generate a random id to replace student numbers
        map[regexesci(id)] = generateRandomId()

        # First, we add the first and last name as delimited in the CSV, both in
        # normal version and in a version without accents. If the name is longer
        # than 4 characters, we just match it anywhere. If it shorter, a
        # non-letter is required before and after, to avoid things like replacing
        # "Han" in every *Handler class.
        for n in name.split(","):
            n = n.strip()
            if not n.lower() in skipNames:
                if len(n) > 4:
                    map[regexesci(n)] = generateRandomName()
                    map[regexesci(removeAccents(n))] = map[regexesci(n)]
                else:
                    map[regexi(r"(^|[^a-zA-Z])" + re.escape(n) + r"([^a-zA-Z]|$)")] = \
                        lambda match: (match.group(1) + map[regexesci(n)] + match.group(2))
                    map[regexi(r"(^|[^a-zA-Z])" + re.escape(removeAccents(n)) + r"([^a-zA-Z]|$)")] = \
                        lambda match: (match.group(1) + map[regexesci(n)] + match.group(2))

        # Then we split the name by spaces, because people may not use their
        # full name. Parts with length 4 or shorter are filtered out to avoid
        # replacing every "de", "van" and "het". (Especially the first one is
        # quite important for .pde files.) If a full name is shorter than 4
        # characters, it is covered in the previous step.
        for namePart in [ namePart.strip() for n in name.split(",") for namePart in n.split(" ") ]:
            if len(namePart) > 4 and namePart.lower() not in skipNames:
                map[regexesci(namePart)] = generateRandomName()
                map[regexesci(removeAccents(namePart))] = map[regexesci(namePart)]

    return map

# Read the students form the gradebook CSV export, returning a dictionary with
# the Canvas user id as key and a tuple with their name and student number as
# value
def readStudentsFromGradebookCsv(gradebookCsvPath: str) -> Dict[str, Tuple[str, str]]:
    with open(gradebookCsvPath, mode = "r", encoding="utf-8") as gradebookFile:
        reader = csv.reader(gradebookFile)
        return {
            # Column 1 is the user id in Canvas
            # Column 0 is the name in Last, First format, column 2 is the student number
            row[1]: (row[0], row[2]) 
            for row in reader 
            # Strip out the double header and the test user
            if row[0].strip() not in [ "Student", "Points Possible", "student, Test", "" ]
        }

# Read assignment information, rubrics and assessments from a JSON file, which
# is created from a GrahpQL query (which can be executed in GrahpiQL)
# query AssignmentRubrics {
#   assignment(id: "...") {
#     name
#     course {
#       name
#     }
#     rubric {
#       title
#       freeFormCriterionComments
#       criteria {
#         _id
#         description
#         longDescription
#         points
#         ratings {
#           _id
#           description
#           longDescription
#           points
#         }
#       }
#     }
#     submissionsConnection(filter: {gradingStatus: graded}) {
#       nodes {
#         attachments {
#           _id
#         }
#         score
#         rubricAssessmentsConnection {
#           nodes {
#             assessmentRatings {
#               _id
#               comments
#               points
#               criterion {
#                 _id
#               }
#             }
#           }
#         }
#       }
#     }
#   }
# }
def readCanvasRubricsJson(canvasRubricsJsonPath: str) -> Tuple[dict, List[dict]]:
    with open(canvasRubricsJsonPath, mode = "r", encoding = "utf-8") as file:
        data = json.load(file)
    assignment: dict = data["data"]["assignment"]
    assessments: List[dict] = assignment["submissionsConnection"]["nodes"]
    return assignment, assessments

# Copy code files from the source zip file to the destination zip file,
# including .jars and traversing nested .zip files. Replacements are applied in
# code files en filenames
def copyZipContent(src: ZipFile, dest: ZipFile, basePath: str, replacements: Replacements, verbose: bool = False):
    # Get all files from the zip, skipping those in a __MACOSX folder (who
    # thought that thing was a good idea?)
    infolist = [ info for info in src.infolist() if "__MACOSX" not in info.filename ]
    for info in infolist:
        ext = path.splitext(info.filename)[1]
        newFilename = replaceAll(info.filename, replacements)

        try:
            # Copy files with .pde, .java and .xml (for Maven projects)
            # extensions, replacing names, student numbers and emails
            if ext in [ ".pde", ".java", ".xml" ]:
                if verbose: print(f"Copying {info.filename} to {path.join(basePath, newFilename)}")
                with src.open(info, mode = "r") as file:
                    fileText = file.read().decode("utf-8")
                fileText = replaceAll(fileText, replacements)
                fileText = replaceEmails(fileText)
                dest.writestr(path.join(basePath, newFilename), fileText.encode("utf-8"))

            # Copy .jar files without replacements, because these are binary files.
            # Useful for code analysis if these are external denpendencies used in
            # the project.
            elif ext in [ ".jar" ]:
                if verbose: print(f"Copying {info.filename} to {path.join(basePath, newFilename)}")
                with src.open(info, mode = "r") as file:
                    dest.writestr(path.join(basePath, newFilename), file.read())

            # Recursively copy from nested zip files, because this apparently
            # happens...
            elif ext == ".zip":
                if verbose: print(f"Found .zip file {info.filename}, starting recursive copy")
                subPath = path.splitext(newFilename)[0]
                with src.open(info, mode = "r") as file:
                    with ZipFile(file, mode = "r") as zipFile:
                        copyZipContent(zipFile, dest, path.join(basePath, subPath), replacements, verbose = verbose)
                if verbose: print(f"Finished recursive copy of {info.filename}")
    
        except UnicodeDecodeError as err:
            print(f"  Failed to decode {ext} file, skipping...")
            if verbose: print(err)
        except Exception as err:
            print(f"  Failed to copy {ext} file, skipping...")
            if verbose: print(err)

# Copy every attachment from the Canvas export (submissions.zip) to the output zip file
def copySubmissions(attachmentsZip: ZipFile, outputZip: ZipFile, copyAssessment: Callable[[str, str, str, ZipFile, str], None], replacements: Replacements, verbose: bool = False):
    # Count the submissions to give a progress indicator
    attachmentCount = len(attachmentsZip.infolist())

    # Loop over all submissions in randomized order
    for i, attachmentInfo in enumerate(shuffled(attachmentsZip.infolist())):
        print(f"Copying submission {i}/{attachmentCount}" + (f": {attachmentInfo.filename}" if verbose else ""))

        newId = str(i).zfill(3)
        basePath = "submissions/" + newId + "/"

        # Find the attachment id from the filename and look up the
        # assessment for the submission with this attachment
        filenameParts = attachmentInfo.filename.split("_")
        if filenameParts[1] == "LATE":
            userId = filenameParts[2]
            attachmentId = filenameParts[3]
        else:
            userId = filenameParts[1]
            attachmentId = filenameParts[2]
        copyAssessment(userId, attachmentId, newId, outputZip, basePath)

        try:
            # Open the inner zip file containing the submitted files
            with attachmentsZip.open(attachmentInfo, mode = "r") as attachmentFile:
                with ZipFile(attachmentFile, mode = "r") as attachmentZip:
                    copyZipContent(attachmentZip, outputZip, basePath, replacements, verbose = verbose)
        except Exception as err:
            print("  Failed to copy submission, skipping...")
            if verbose: print(err)

# Copy rubric-based assessments related to a certain attachment from the Canvas
# JSON export to the output zip file at the specified path
def copyRubricAssessment(rubricAssessments: List[dict], attachmentId: str, outputZip: ZipFile, basePath: str, replacements: Replacements, verbose: bool = False):
    assessment = [ 
        { "score": s["score"], "rubrics": s["rubricAssessmentsConnection"]["nodes"] }
        for s in rubricAssessments
        if any([ a["_id"] == attachmentId for a in s["attachments"] ])
    ]
    try:
        outputZip.writestr(
            path.join(basePath, "assessment.json"), 
            json.dumps(replaceInObject(assessment, replacements), indent = 2)
        )
    except Exception as err:
        print("  Failed to write assessment, skipping...")
        if verbose: print(err)

def renameGradingSheet(gradingWorkbook: Workbook, studentNr: str, newId: str, verbose: bool = False):
    for sheet in gradingWorkbook:
        if (sheet["E3"].value is not None and studentNr in sheet["E3"].value) or \
           (sheet["E4"].value is not None and studentNr in sheet["E4"].value):
            if sheet.title.startswith("n"):
                newTitle = "n" + newId + ", " + sheet.title
            else:
                newTitle = "n" + newId + " ".join(sheet.title.split(" ")[1:])
            if sheet.title.startswith("Resit"):
                newTitle += " (resit)"
            if verbose: print(f"Renaming grading sheet {sheet.title} to {newTitle}")
            sheet.title = newTitle

def main(attachmentsZipPath: str, gradebookCsvPath: str, canvasRubricsJsonPath: Union[str, None] = None, gradingSheetsPath: Union[str, None] = None, verbose: bool = False):
    # Determine the path of the output .zip file
    outputZipPath = path.splitext(attachmentsZipPath)[0] + "_cleaned.zip"
    
    # Basic checks on the input to see if all is valid
    errors = []
    if not zipfile.is_zipfile(attachmentsZipPath):
        errors += [ f"{attachmentsZipPath} is not a valid zip file." ]
    if path.exists(outputZipPath):
        errors += [ f"{outputZipPath} already exists, please remove it first." ]
    if not path.exists(gradebookCsvPath):
        errors += [ f"{gradebookCsvPath} does not exist." ]
    if canvasRubricsJsonPath is not None and not path.exists(canvasRubricsJsonPath):
        errors += [ f"{canvasRubricsJsonPath} does not exist." ]
    if gradingSheetsPath is not None and not path.exists(gradingSheetsPath):
        errors += [ f"{gradingSheetsPath} does not exist." ]
    if len(errors) > 0:
        for error in errors:
            print(error)
        sys.exit(1)

    # Read list of students and generate replacements for PII
    students = readStudentsFromGradebookCsv(gradebookCsvPath)
    replacements = generatePiiReplacements(students.values())

    # Read rubrics and assessments from a configured location
    if canvasRubricsJsonPath is not None:
        assignment, assessments = readCanvasRubricsJson(canvasRubricsJsonPath)
        copyAssessment = lambda userId, attachmentId, newId, outputZip, basePath: \
            copyRubricAssessment(assessments, attachmentId, outputZip, basePath, replacements, verbose = verbose)
    elif gradingSheetsPath is not None:
        assignment = None
        print("Loading grading sheets...")
        assessments = load_workbook(filename = gradingSheetsPath)
        copyAssessment = lambda userId, attachmentId, newId, outputZip, basePath: \
            renameGradingSheet(assessments, students[userId][1], newId, verbose = verbose)
    else:
        assignment, assessments = None, None
        copyAssessment = lambda userId, attachmentId, newId, outputZip, basePath: None

    # Write the output .zip file
    with ZipFile(outputZipPath, mode = "x", compression = zipfile.ZIP_DEFLATED) as outputZip:
        if assignment is not None:
            print("Writing assignment information and rubric...")
            outputZip.writestr(
                "info.json", 
                json.dumps({ "name": assignment["name"], "courseName": assignment["course"]["name"] }, indent = 2)
            )
            outputZip.writestr("rubric.json", json.dumps(assignment["rubric"], indent = 2))

        # Copy submissions from the source zip file
        with ZipFile(attachmentsZipPath, mode = "r") as attachmentsZip:
            copySubmissions(attachmentsZip, outputZip, copyAssessment, replacements, verbose = verbose)

        if gradingSheetsPath is not None:
            # Remove all grading sheets for which we don't have a matching
            # submission (which weren't renamed)
            for sheet in assessments:
                if not sheet.title.startswith("n"):
                    assessments.remove(sheet)

            # Sort the grading sheets by title
            for i, sheet in enumerate(sorted(assessments.sheetnames)):
                assessments.move_sheet(sheet, i)
            
            # Set the first sheet as active
            assessments.active = 0

            # Clear the cells with group name and student numbers
            for sheet in assessments:
                sheet["E2"].value = None
                sheet["E3"].value = None
                sheet["E4"].value = None

            print("Copying grading sheets...")
            # Save the grading sheets and add to the output zip file
            tmpGradingSheetsPath = path.splitext(gradingSheetsPath)[0] + "_tmp.xlsx"
            assessments.save(tmpGradingSheetsPath)
            outputZip.write(tmpGradingSheetsPath, "gradingSheets.xlsx")
            remove(tmpGradingSheetsPath)

    print("Done.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "-v", "--verbose", 
        action = "store_true", 
        help = "Print more diagnostic information. Note: this prints information that will allow deanonymization of the output!"
    )

    parser.add_argument("submissions", help = "Path to the submissions.zip export from Canvas")
    parser.add_argument("gradebook", help = "Path to the gradebook CSV export from Canvas")

    parser.add_argument(
        "-r", "--rubrics", 
        help = "Path to the JSON file containing rubrics and assessments, exported from Canvas via GraphiQL")
    parser.add_argument(
        "-g", "--gradingsheets",
        help = "Path to the Excel workbook containing grading sheets in SS format."
    )

    args = parser.parse_args()

    main(args.submissions, args.gradebook, args.rubrics, args.gradingsheets, verbose = args.verbose)
